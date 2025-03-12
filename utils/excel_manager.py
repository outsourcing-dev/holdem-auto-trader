from utils.encrypt_excel import EncryptExcel, decrypt_auto_excel
import os
import logging
import time
import openpyxl
from typing import Dict, Any, Optional, Tuple
import atexit
import sys

# 로거 설정
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 암호화된 Excel 사용 여부 (True면 암호화 기능 사용)
USE_ENCRYPTED_EXCEL = True
EXCEL_PASSWORD = "holdem2025"  # 기본 암호

# Windows에서만 사용 가능한 COM 인터페이스
try:
    import win32com.client
    import pythoncom
    HAS_WIN32COM = True
    logger.info("Excel COM 인터페이스 사용 가능")
except ImportError:
    HAS_WIN32COM = False
    logger.info("Excel COM 인터페이스 사용 불가 - win32com 라이브러리 없음")

class ExcelManager:
    def __init__(self, excel_path: str = None):
        """
        엑셀 파일 관리자 초기화
        
        Args:
            excel_path (str): 엑셀 파일 경로 (None이면 자동 탐지)
        """
        # Excel 관련 속성 초기화
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        
        # Excel 관련 속성 초기화
        self.excel_app = None
        self.workbook = None
        self.is_excel_open = False
        
        # 명시적으로 경로 지정된 경우
        if excel_path is not None:
            self.excel_path = excel_path
            logger.info(f"지정된 Excel 파일 경로 사용: {self.excel_path}")
        else:
            # 환경 변수에서 경로 가져오기 시도
            env_path = os.environ.get("AUTO_EXCEL_PATH")
            if env_path and os.path.exists(env_path):
                self.excel_path = env_path
                logger.info(f"환경 변수에서 Excel 파일 경로 가져옴: {self.excel_path}")
            else:
                # 환경 변수 없을 경우 main.py의 전역 변수 확인
                try:
                    import sys
                    main_module = sys.modules.get('__main__')
                    if main_module and hasattr(main_module, 'global_excel_path'):
                        self.excel_path = main_module.global_excel_path
                        logger.info(f"전역 변수에서 Excel 파일 경로 가져옴: {self.excel_path}")
                    else:
                        # 마지막으로 AUTO.xlsx, AUTO.encrypted 확인
                        self.excel_path = self.get_excel_path("AUTO.xlsx")
                        
                        if not os.path.exists(self.excel_path):
                            # AUTO.encrypted 파일 경로 확인
                            encrypted_path = self.get_excel_path("AUTO.encrypted")
                            if os.path.exists(encrypted_path):
                                # 임시 파일 생성
                                temp_dir = tempfile.gettempdir()
                                temp_excel_path = os.path.join(temp_dir, f"AUTO_temp_{os.getpid()}.xlsx")
                                
                                # 복호화 시도
                                encryptor = EncryptExcel()
                                if encryptor.decrypt_file(encrypted_path, temp_excel_path, "holdem2025_secret_key"):
                                    self.excel_path = temp_excel_path
                                    logger.info(f"AUTO.encrypted 파일 자체 복호화 성공: {self.excel_path}")
                                else:
                                    logger.error(f"AUTO.encrypted 파일 자체 복호화 실패")
                except Exception as e:
                    logger.error(f"Excel 파일 경로 자동 감지 중 오류: {e}")
                    self.excel_path = self.get_excel_path("AUTO.xlsx")
                    
        # Excel 암호화 관리자
        self.encryptor = EncryptExcel()
        
        # 프로그램 종료 시 Excel 종료 보장
        atexit.register(self.close_excel)
        
        # 암호화된 Excel 파일 확인 및 복호화
        self._check_and_decrypt_excel()
        
        # 엑셀 파일이 존재하는지 확인
        if not os.path.exists(self.excel_path):
            logger.error(f"엑셀 파일을 찾을 수 없습니다: {self.excel_path}")
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {self.excel_path}")
        
        logger.info(f"Excel 매니저 초기화 완료 - 파일 경로: {self.excel_path}")
        
        # 프로그램 시작 시 Excel 열기 시도
        if HAS_WIN32COM:
            self.open_excel_once()

    # utils/excel_manager.py의 get_excel_path 메서드 수정
    def get_excel_path(self, filename="AUTO.xlsx"):
        """
        실행 환경에 따라 Excel 파일의 적절한 경로를 반환합니다.
        """
        # 암호화된 파일을 찾는 경우 파일명 수정
        if filename == "AUTO.xlsx.enc":
            filename = "AUTO.encrypted"  # 이름 통일
            
        if getattr(sys, 'frozen', False):
            # PyInstaller로 빌드된 실행 파일인 경우
            base_dir = os.path.dirname(sys.executable)
            excel_path = os.path.join(base_dir, filename)
            logger.info(f"PyInstaller 환경에서 Excel 파일 경로: {excel_path}")
        else:
            # 일반 Python 스크립트로 실행되는 경우
            excel_path = filename
            logger.info(f"일반 Python 환경에서 Excel 파일 경로: {excel_path}")
        
        # 파일 존재 여부 확인
        if not os.path.exists(excel_path):
            logger.warning(f"Excel 파일을 찾을 수 없습니다: {excel_path}")
        
        return excel_path

    def _check_and_decrypt_excel(self):
        """암호화된 Excel 파일이 있는지 확인하고 필요시 복호화"""
        if not USE_ENCRYPTED_EXCEL:
            return
            
        # 암호화된 파일 경로
        encrypted_excel_path = self.get_excel_path("AUTO.xlsx.enc")
        
        # 암호화된 파일이 있고 일반 파일이 없으면 복호화
        if os.path.exists(encrypted_excel_path) and not os.path.exists(self.excel_path):
            logger.info(f"암호화된 Excel 파일 발견: {encrypted_excel_path}, 복호화 시도...")
            
            # 복호화 시도
            if decrypt_auto_excel(EXCEL_PASSWORD):
                logger.info("Excel 파일 복호화 완료")
            else:
                logger.error("Excel 파일 복호화 실패")
                raise FileNotFoundError("Excel 파일을 복호화할 수 없습니다")

    # 이하 기존 메서드들...
    
    def __del__(self):
        """객체 소멸 시 Excel 종료 보장"""
        self.close_excel()
    
    def open_excel_once(self):
        """Excel 애플리케이션을 한 번 열고 계속 사용"""
        if not HAS_WIN32COM:
            return False
            
        # 기존에 열려있으면 먼저 닫기
        if self.is_excel_open:
            self.close_excel()
            time.sleep(0.5)  # 잠시 대기
        
        try:
            # COM 스레드 초기화
            pythoncom.CoInitialize()
            
            # Excel 애플리케이션 실행
            self.excel_app = win32com.client.Dispatch("Excel.Application")
            self.excel_app.Visible = False
            self.excel_app.DisplayAlerts = False
            
            # 절대 경로로 변환
            abs_path = os.path.abspath(self.excel_path)
            
            # 워크북 열기
            self.workbook = self.excel_app.Workbooks.Open(abs_path)
            
            self.is_excel_open = True
            logger.info("Excel 애플리케이션 시작 및 파일 로드 완료")
            return True
        except Exception as e:
            logger.error(f"Excel 애플리케이션 시작 실패: {e}")
            self.close_excel()
            return False
        
    def close_excel(self):
        """Excel 애플리케이션 종료"""
        try:
            if self.workbook:
                try:
                    self.workbook.Save()
                    logger.info("Excel 파일 저장 완료")
                except Exception as e:
                    logger.warning(f"Excel 저장 중 오류: {e}")
                
                try:
                    self.workbook.Close(True)  # True: 변경 사항 저장
                    logger.info("Excel 워크북 닫기 완료")
                except Exception as e:
                    logger.warning(f"Excel 닫기 중 오류: {e}")
                
                self.workbook = None
            
            if self.excel_app:
                try:
                    self.excel_app.Quit()
                    logger.info("Excel 애플리케이션 종료 완료")
                except Exception as e:
                    logger.warning(f"Excel 종료 중 오류: {e}")
                
                self.excel_app = None
            
            # COM 스레드 해제
            try:
                pythoncom.CoUninitialize()
            except:
                pass
            
            self.is_excel_open = False
            
            # 종료 시 자동 암호화 
            if USE_ENCRYPTED_EXCEL:
                # 종료 후 원본 파일 암호화 (기존 암호화 파일 덮어쓰기)
                encrypted_path = self.get_excel_path("AUTO.xlsx.enc")
                self.encryptor.encrypt_file(self.excel_path, encrypted_path, EXCEL_PASSWORD)
                logger.info(f"Excel 파일 '{self.excel_path}'를 종료 시 자동 암호화 완료")
        except Exception as e:
            logger.warning(f"Excel 종료 중 오류: {e}")
    
    def reopen_excel_if_needed(self):
        """필요한 경우 Excel을 다시 엽니다"""
        if not self.is_excel_open or not self.excel_app or not self.workbook:
            logger.info("Excel 재연결이 필요하여 다시 여는 중...")
            return self.open_excel_once()
        return True
        
    def update_formulas(self):
        """수식 업데이트만 수행 (저장 없이)"""
        if not self.reopen_excel_if_needed():
            return False
            
        try:
            self.workbook.Application.Calculate()
            return True
        except Exception as e:
            logger.error(f"수식 업데이트 실패: {e}")
            self.reopen_excel_if_needed()
            return False
        
    def save_without_close(self):
        """파일 저장 (닫지 않고)"""
        if not self.reopen_excel_if_needed():
            return False
            
        try:
            start_time = time.time()
            self.workbook.Save()
            elapsed = time.time() - start_time
            logger.info(f"Excel 파일 저장 완료 (소요시간: {elapsed:.2f}초)")
            return True
        except Exception as e:
            logger.error(f"파일 저장 실패: {e}")
            self.reopen_excel_if_needed()
            return False
    
    def read_row(self, row_number: int, start_col: str = 'B', end_col: str = 'R') -> Dict[str, Any]:
        """
        특정 행의 값을 읽어옵니다.
        
        Args:
            row_number (int): 읽어올 행 번호
            start_col (str): 시작 열 문자 (기본값: 'B')
            end_col (str): 끝 열 문자 (기본값: 'R')
        
        Returns:
            Dict[str, Any]: 열 이름을 키로 하고 셀 값을 값으로 하는 딕셔너리
        """
        # COM 인스턴스를 사용하여 읽기 시도
        if not self.reopen_excel_if_needed():
            return {}
            
        try:
            # 수식 업데이트
            self.update_formulas()
            
            # 결과를 저장할 딕셔너리 초기화
            result = {}
            
            # 열 문자를 열 인덱스로 변환
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)
            
            sheet = self.workbook.ActiveSheet
            
            # 지정된 행의 열별로 값 읽기
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                cell_value = sheet.Cells(row_number, col_idx).Value
                result[col_letter] = cell_value
            
            return result
        except Exception as e:
            logger.error(f"COM 인터페이스로 행 읽기 실패: {e}")
            self.reopen_excel_if_needed()
            return {}
    
    def read_pick_row(self) -> Dict[str, str]:
        """
        12행(PICK 행)의 값을 읽어옵니다.
        
        Returns:
            Dict[str, str]: 열 이름을 키로 하고 셀 값을 값으로 하는 딕셔너리
        """
        row_data = self.read_row(12)
        
        # 문자열로 변환
        for col, value in row_data.items():
            if value is None:
                row_data[col] = "N"  # None은 'N'으로 처리
            elif not isinstance(value, str):
                row_data[col] = str(value)
        
        return row_data
    
    def read_result_row(self) -> Dict[str, str]:
        """
        16행(결과 행)의 값을 읽어옵니다.
        
        Returns:
            Dict[str, str]: 열 이름을 키로 하고 셀 값(W/L/N)을 값으로 하는 딕셔너리
        """
        row_data = self.read_row(16)
        
        # 문자열로 변환
        for col, value in row_data.items():
            if value is None:
                row_data[col] = "N"  # None은 'N'으로 처리
            elif not isinstance(value, str):
                row_data[col] = str(value)
        
        return row_data
    
    def read_cell_value(self, column, row):
        """특정 셀 값 읽기 (COM 인터페이스 사용)"""
        if not self.reopen_excel_if_needed():
            return None
            
        try:
            sheet = self.workbook.ActiveSheet
            col_idx = openpyxl.utils.column_index_from_string(column)
            return sheet.Cells(row, col_idx).Value
        except Exception as e:
            logger.error(f"셀 값 읽기 실패: {e}")
            self.reopen_excel_if_needed()
            return None
    
    def write_game_result(self, column: str, result: str) -> bool:
        """
        게임 결과를 3행에 씁니다.
        COM 인스턴스 재사용 방식을 사용합니다.
        
        Args:
            column (str): 열 문자 (예: 'B', 'C', 'D')
            result (str): 결과 값 ('P', 'B', 'T' 중 하나)
        
        Returns:
            bool: 성공 여부
        """
        if not self.reopen_excel_if_needed():
            return False
            
        try:
            start_time = time.time()
            sheet = self.workbook.ActiveSheet
            
            # 3행에 결과 쓰기
            col_idx = openpyxl.utils.column_index_from_string(column)
            sheet.Cells(3, col_idx).Value = result
            
            # 저장 (닫지 않고)
            self.save_without_close()
            
            # 수식 업데이트
            self.update_formulas()
            
            elapsed = time.time() - start_time
            logger.info(f"{column}3에 '{result}' 기록 완료 (소요시간: {elapsed:.2f}초)")
            return True
        except Exception as e:
            logger.error(f"엑셀 파일에 게임 결과 쓰기 실패: {e}")
            self.reopen_excel_if_needed()
            return False
    
    def get_next_empty_column(self, row: int = 3, start_col: str = 'B', end_col: str = 'BW') -> Optional[str]:
        """
        지정된 행에서 값이 비어 있는 첫 번째 열을 찾습니다.
        
        Args:
            row (int): 확인할 행 번호 (기본값: 3)
            start_col (str): 시작 열 문자 (기본값: 'B')
            end_col (str): 끝 열 문자 (기본값: 'BW')
        
        Returns:
            Optional[str]: 값이 비어 있는 첫 번째 열 문자 또는 None (모든 열이 채워진 경우)
        """
        # COM 인스턴스 사용 시도
        if not self.reopen_excel_if_needed():
            return None
            
        try:
            sheet = self.workbook.ActiveSheet
            
            # 열 인덱스 변환
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)
            
            # 비어 있는 열 찾기
            for col_idx in range(start_col_idx, end_col_idx + 1):
                cell_value = sheet.Cells(row, col_idx).Value
                if cell_value is None or cell_value == "":
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    return col_letter
            
            return None
        except Exception as e:
            logger.error(f"COM으로 빈 열 찾기 실패: {e}")
            self.reopen_excel_if_needed()
            return None
    
    def get_current_column(self) -> Optional[str]:
        """
        현재 작업 중인 열을 찾습니다. (3행 기준으로 비어 있는 첫 번째 열)
        
        Returns:
            Optional[str]: 현재 작업 중인 열 문자 또는 None
        """
        return self.get_next_empty_column(row=3)
    
    def check_betting_needed(self, column: str) -> Tuple[bool, str]:
        """
        지정된 열의 12행 값을 확인하여 배팅이 필요한지 결정합니다.
        
        Args:
            column (str): 열 문자 (예: 'B', 'C', 'D')
        
        Returns:
            Tuple[bool, str]: (배팅 필요 여부, PICK 값)
            - 배팅 필요 여부: PICK 값이 'B' 또는 'P'일 때 True, 그 외에는 False
            - PICK 값: 'B'(뱅커), 'P'(플레이어), 'N'(배팅 안 함) 또는 다른 값
        """
        if not self.reopen_excel_if_needed():
            return (False, 'N')
            
        try:
            # 수식 업데이트
            self.update_formulas()
            
            # 셀 값 읽기
            pick_value = self.read_cell_value(column, 12)
            
            # 값이 None이면 'N'으로 처리
            if pick_value is None:
                pick_value = 'N'
            # 문자열이 아니면 문자열로 변환
            elif not isinstance(pick_value, str):
                pick_value = str(pick_value)
            
            # 배팅 필요 여부 결정
            need_betting = pick_value in ['B', 'P']
            return (need_betting, pick_value)
        except Exception as e:
            logger.error(f"PICK 값 확인 실패: {e}")
            self.reopen_excel_if_needed()
            return (False, 'N')  # 오류 발생 시 배팅 안 함
    
    def check_result(self, column: str) -> Tuple[bool, str]:
        """
        지정된 열의 16행 값을 확인하여 승패 결과를 확인합니다.
        
        Args:
            column (str): 열 문자 (예: 'B', 'C', 'D')
        
        Returns:
            Tuple[bool, str]: (성공 여부, 결과 값)
            - 성공 여부: 결과 값이 'W'일 때 True, 그 외에는 False
            - 결과 값: 'W'(승리), 'L'(패배), 'N'(미배팅) 또는 다른 값
        """
        if not self.reopen_excel_if_needed():
            return (False, 'N')
            
        try:
            # 수식 업데이트
            self.update_formulas()
            
            # 셀 값 읽기
            result_value = self.read_cell_value(column, 16)
            
            # 값이 None이면 'N'으로 처리
            if result_value is None:
                result_value = 'N'
            # 문자열이 아니면 문자열로 변환
            elif not isinstance(result_value, str):
                result_value = str(result_value)
            
            # 성공 여부 결정
            is_success = result_value == 'W'
            return (is_success, result_value)
        except Exception as e:
            logger.error(f"결과 값 확인 실패: {e}")
            self.reopen_excel_if_needed()
            return (False, 'N')  # 오류 발생 시 미배팅으로 처리
    
    def get_current_round_info(self) -> Dict[str, Any]:
        """
        현재 라운드 정보를 반환합니다.
        - 다음에 결과를 입력할 열
        - 다음 라운드에 배팅해야 하는지 여부
        - PICK 값
        
        Returns:
            Dict[str, Any]: 현재 라운드 정보
        """
        # 현재 열 찾기
        current_column = self.get_current_column()
        
        if current_column is None:
            return {
                "round_column": None,
                "need_betting": False,
                "pick_value": 'N',
                "message": "모든 열이 채워져 있습니다."
            }
        
        # 배팅 필요 여부 확인
        need_betting, pick_value = self.check_betting_needed(current_column)
        
        return {
            "round_column": current_column,
            "need_betting": need_betting,
            "pick_value": pick_value,
            "message": f"{current_column} 열, PICK 값: {pick_value}"
        }
    
    def clear_row(self, row_number: int, start_col: str = 'B', end_col: str = 'BW') -> bool:
        """
        지정된 행의 값을 모두 지웁니다.
        
        Args:
            row_number (int): 지울 행 번호
            start_col (str): 시작 열 문자 (기본값: 'B')
            end_col (str): 끝 열 문자 (기본값: 'BW')
        
        Returns:
            bool: 성공 여부
        """
        if not self.reopen_excel_if_needed():
            return False
            
        try:
            sheet = self.workbook.ActiveSheet
            
            # 열 인덱스 변환
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)
            
            # 범위 지정 초기화
            clear_range = sheet.Range(
                sheet.Cells(row_number, start_col_idx), 
                sheet.Cells(row_number, end_col_idx)
            )
            clear_range.ClearContents()
            
            # 저장
            self.save_without_close()
            
            logger.info(f"{row_number}행 {start_col}~{end_col} 열 초기화 완료 (COM)")
            return True
        except Exception as e:
            logger.error(f"{row_number}행 초기화 실패: {e}")
            self.reopen_excel_if_needed()
            return False

    def initialize_excel(self) -> bool:
        """
        엑셀 파일 초기화
        - 3행(결과 행) 초기화
        
        Returns:
            bool: 성공 여부
        """
        # 3행 초기화
        return self.clear_row(3, 'B', 'BW')
    
    def write_game_results_sequence(self, results):
        """
        게임 결과 시퀀스를 B3부터 순서대로 엑셀에 기록합니다.
        
        Args:
            results (list): 게임 결과 리스트 (예: ['P', 'B', 'T', 'B', ...])
            
        Returns:
            bool: 성공 여부
        """
        if not self.reopen_excel_if_needed():
            logger.error("Excel에 연결할 수 없어 게임 결과를 기록할 수 없습니다.")
            return False
            
        try:
            # COM 인스턴스를 사용하여 기록
            sheet = self.workbook.ActiveSheet
            
            # 3행 전체를 초기화
            start_col = 2  # B열(인덱스 2)
            end_col = 75   # BW열(인덱스 75)
            
            logger.info("3행 초기화 중...")
            clear_range = sheet.Range(
                sheet.Cells(3, start_col), 
                sheet.Cells(3, end_col)
            )
            clear_range.ClearContents()
            logger.info("3행 초기화 완료")
            
            # 결과 시퀀스 기록
            for idx, result in enumerate(results):
                col_idx = 2 + idx  # B(2)부터 시작
                sheet.Cells(3, col_idx).Value = result
            
            # 저장 및 수식 업데이트
            self.workbook.Save()
            self.update_formulas()
            
            logger.info(f"총 {len(results)}개의 결과를 Excel COM으로 기록 완료")
            return True
        except Exception as e:
            logger.error(f"게임 결과 시퀀스 기록 중 오류 발생: {e}")
            # 에러 발생 시 재연결 시도
            self.reopen_excel_if_needed()
            return False
        
    def get_next_column_letter(self, column):
        """
        지정된 열의 다음 열 문자를 반환합니다.
        
        Args:
            column (str): 열 문자 (예: 'B')
            
        Returns:
            str: 다음 열 문자 (예: 'C')
        """
        try:
            col_idx = openpyxl.utils.column_index_from_string(column)
            return openpyxl.utils.get_column_letter(col_idx + 1)
        except Exception as e:
            logger.error(f"다음 열 문자 가져오기 실패: {e}")
            return None
    
    def get_prev_column_letter(self, column):
        """
        지정된 열의 이전 열 문자를 반환합니다.
        
        Args:
            column (str): 열 문자 (예: 'C')
            
        Returns:
            str: 이전 열 문자 (예: 'B') 또는 열이 'A'인 경우 None
        """
        try:
            col_idx = openpyxl.utils.column_index_from_string(column)
            if col_idx <= 1:  # 'A' 이하면 이전 열이 없음
                return None
            return openpyxl.utils.get_column_letter(col_idx - 1)
        except Exception as e:
            logger.error(f"이전 열 문자 가져오기 실패: {e}")
            return None
    
    def check_next_column_pick(self, last_result_column):
        """
        마지막으로 결과가 입력된 열의 다음 열에서 12행(PICK) 값을 확인합니다.
        
        Args:
            last_result_column (str): 마지막 결과가 입력된 열 문자 (예: 'J')
            
        Returns:
            str: 다음 열의 PICK 값 ('P', 'B', 'N' 중 하나)
        """
        if not self.reopen_excel_if_needed():
            self.logger.warning("Excel 연결을 재설정할 수 없습니다. 기본값 'N' 반환")
            return 'N'
            
        try:
            # 열 인덱스 계산
            col_idx = openpyxl.utils.column_index_from_string(last_result_column)
            next_col_idx = col_idx + 1
            next_col_letter = openpyxl.utils.get_column_letter(next_col_idx)
            
            # 수식 업데이트 - 중요: 매번 수식을 재계산하도록 함
            self.update_formulas()
            time.sleep(0.2)  # 약간의 대기 시간 추가
            
            # 12행 값 읽기
            pick_value = self.read_cell_value(next_col_letter, 12)
            
            # 값이 None이거나 빈 문자열이면 'N'으로 처리
            if pick_value is None or pick_value == "":
                # 추가 검증: Excel 파일에서 다시 직접 읽어보기
                try:
                    # 다시 한번 시도
                    workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
                    sheet = workbook.active
                    direct_value = sheet[f"{next_col_letter}12"].value
                    workbook.close()
                    
                    # 직접 읽은 값이 있으면 사용
                    if direct_value not in [None, ""]:
                        self.logger.info(f"COM 인터페이스에서 빈 값이 반환되었지만 직접 읽기에서 값 발견: {direct_value}")
                        return str(direct_value)
                except Exception as e:
                    self.logger.warning(f"직접 읽기 시도 중 오류: {e}")
                
                # 마지막 시도: 강제로 수식 계산 후 다시 시도
                try:
                    if HAS_WIN32COM:
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False
                        wb = excel.Workbooks.Open(os.path.abspath(self.excel_path))
                        excel.Calculate()  # 모든 수식 계산
                        sheet = wb.Sheets(1)
                        forced_value = sheet.Cells(12, next_col_idx).Value
                        wb.Close(False)
                        excel.Quit()
                        
                        if forced_value not in [None, ""]:
                            self.logger.info(f"강제 수식 계산 후 값 발견: {forced_value}")
                            return str(forced_value)
                except Exception as e:
                    self.logger.warning(f"강제 수식 계산 시도 중 오류: {e}")
                
                self.logger.warning(f"다음 열 {next_col_letter}12의 PICK 값이 비어있거나 없음. 'N' 반환")
                return 'N'
                
            # 문자열이 아니면 문자열로 변환
            elif not isinstance(pick_value, str):
                pick_value = str(pick_value)
            
            # PICK 값이 빈 문자열이면 'N'으로 처리
            if pick_value.strip() == "":
                self.logger.warning(f"다음 열 {next_col_letter}12의 PICK 값이 빈 문자열. 'N' 반환")
                return 'N'
                
            # Pick 값 검증 - P, B 외의 값이 나오면 경고 로그
            if pick_value not in ['P', 'B', 'N']:
                self.logger.warning(f"예상치 못한 PICK 값: {pick_value}, 엑셀 파일 검증 필요")
            
            start_time = time.time()
            elapsed = time.time() - start_time
            self.logger.info(f"다음 열 {next_col_letter}12의 PICK 값: {pick_value} (소요시간: {elapsed:.2f}초)")
            return pick_value
        except Exception as e:
            self.logger.error(f"다음 열 PICK 값 확인 중 오류 발생: {e}")
            self.reopen_excel_if_needed()
            return 'N'  # 오류 발생 시 기본값 'N' 반환
        
    def write_filtered_game_results(self, filtered_results, actual_results):
        """
        TIE를 포함한 실제 사용 결과를 엑셀에 기록합니다.
        
        Args:
            filtered_results (list): TIE를 제외한 결과 리스트 (P, B만 포함)
            actual_results (list): TIE를 포함한 실제 사용 결과 리스트
            
        Returns:
            bool: 성공 여부
        """
        # actual_results를 엑셀에 기록 (가장 오래된 순서대로)
        return self.write_game_results_sequence(actual_results)