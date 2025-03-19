# utils/excel/formula_manager.py
"""
Excel 수식 관리 클래스
- 수식 계산 및 업데이트
- PICK 값 판단
- 열 단위 작업
"""
import logging
import time
import openpyxl
import os
import re
from typing import Tuple, Any, Optional

class ExcelFormulaManager:
    """Excel 파일의 수식 관련 기능을 제공하는 클래스"""
    
    def __init__(self, base_manager):
        """
        Excel 수식 관리자 초기화
        
        Args:
            base_manager (ExcelBaseManager): 기본 Excel 관리자 객체
        """
        self.base_manager = base_manager
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
    
    def update_formulas(self):
        """
        수식 업데이트만 수행 (저장 없이)
        
        Returns:
            bool: 성공 여부
        """
        if not self.base_manager.reopen_excel_if_needed():
            return False
            
        try:
            self.base_manager.workbook.Application.Calculate()
            return True
        except Exception as e:
            self.logger.error(f"수식 업데이트 실패: {e}")
            self.base_manager.reopen_excel_if_needed()
            return False
    
    def save_without_close(self):
        """
        파일 저장 (닫지 않고)
        
        Returns:
            bool: 성공 여부
        """
        if not self.base_manager.reopen_excel_if_needed():
            return False
            
        try:
            start_time = time.time()
            self.base_manager.workbook.Save()
            elapsed = time.time() - start_time
            return True
        except Exception as e:
            self.logger.error(f"파일 저장 실패: {e}")
            self.base_manager.reopen_excel_if_needed()
            return False
    
    def check_betting_needed(self, column):
        """
        지정된 열의 12행 값을 확인하여 배팅이 필요한지 결정합니다.
        
        Args:
            column (str): 열 문자 (예: 'B', 'C', 'D')
            
        Returns:
            Tuple[bool, str]: (배팅 필요 여부, PICK 값)
        """
        if not self.base_manager.reopen_excel_if_needed():
            return (False, 'N')
            
        try:
            # 수식 업데이트
            self.update_formulas()
            
            # 셀 값 읽기
            pick_value = self._read_cell_value(column, 12)
            
            # 값이 None이면 'N'으로 처리
            if pick_value is None:
                pick_value = 'N'
            # 문자열이 아니면 문자열로 변환
            elif not isinstance(pick_value, str):
                pick_value = str(pick_value)
            
            # 배팅 필요 여부 결정
            need_betting = pick_value in ['B', 'P']
            return (need_betting, pick_value)
        except Exception as e:
            self.logger.error(f"PICK 값 확인 실패: {e}")
            self.base_manager.reopen_excel_if_needed()
            return (False, 'N')  # 오류 발생 시 배팅 안 함
    
    def _read_cell_value(self, column, row):
        """내부 셀 값 읽기 헬퍼 메서드"""
        if not self.base_manager.reopen_excel_if_needed():
            return None
        
        try:
            sheet = self.base_manager.workbook.ActiveSheet
            col_idx = openpyxl.utils.column_index_from_string(column)
            return sheet.Cells(row, col_idx).Value
        except:
            return None
    
    def check_result(self, column):
        """
        지정된 열의 16행 값을 확인하여 승패 결과를 확인합니다.
        
        Args:
            column (str): 열 문자 (예: 'B', 'C', 'D')
            
        Returns:
            Tuple[bool, str]: (성공 여부, 결과 값)
        """
        if not self.base_manager.reopen_excel_if_needed():
            return (False, 'N')
            
        try:
            # 수식 업데이트
            self.update_formulas()
            
            # 셀 값 읽기
            result_value = self._read_cell_value(column, 16)
            
            # 값이 None이면 'N'으로 처리
            if result_value is None:
                result_value = 'N'
            # 문자열이 아니면 문자열로 변환
            elif not isinstance(result_value, str):
                result_value = str(result_value)
            
            # 성공 여부 결정
            is_success = result_value == 'W'
            return (is_success, result_value)
        except Exception as e:
            self.logger.error(f"결과 값 확인 실패: {e}")
            self.base_manager.reopen_excel_if_needed()
            return (False, 'N')  # 오류 발생 시 미배팅으로 처리
    
    def check_next_column_pick(self, last_result_column):
        """
        마지막으로 결과가 입력된 열의 다음 열에서 12행(PICK) 값을 확인합니다.
        
        Args:
            last_result_column (str): 마지막 결과가 입력된 열 문자 (예: 'J')
            
        Returns:
            str: 다음 열의 PICK 값 ('P', 'B', 'N' 중 하나)
        """
        if not self.base_manager.reopen_excel_if_needed():
            return 'N'
            
        try:
            # 열 인덱스 계산
            col_idx = openpyxl.utils.column_index_from_string(last_result_column)
            next_col_idx = col_idx + 1
            next_col_letter = openpyxl.utils.get_column_letter(next_col_idx)
            
            # 수식 업데이트 - 중요: 매번 수식을 재계산하도록 함
            self.update_formulas()
            time.sleep(0.2)  # 약간의 대기 시간 추가
            
            # 12행 값 읽기
            pick_value = self._read_cell_value(next_col_letter, 12)
            
            # 값이 None이거나 빈 문자열이면 'N'으로 처리
            if pick_value is None or pick_value == "":
                # 추가 검증: Excel 파일에서 다시 직접 읽어보기
                try:
                    # 다시 한번 시도
                    workbook = openpyxl.load_workbook(self.base_manager.excel_path, data_only=True)
                    sheet = workbook.active
                    direct_value = sheet[f"{next_col_letter}12"].value
                    workbook.close()
                    
                    # 직접 읽은 값이 있으면 사용
                    if direct_value not in [None, ""]:
                        return str(direct_value)
                except Exception:
                    pass
                
                # 마지막 시도: 강제로 수식 계산 후 다시 시도
                try:
                    if hasattr(self.base_manager, 'excel_app') and self.base_manager.excel_app:
                        sheet = self.base_manager.workbook.Sheets(1)
                        forced_value = sheet.Cells(12, next_col_idx).Value
                        
                        if forced_value not in [None, ""]:
                            return str(forced_value)
                except Exception:
                    pass
                
                return 'N'
                
            # 문자열이 아니면 문자열로 변환
            elif not isinstance(pick_value, str):
                pick_value = str(pick_value)
            
            # PICK 값이 빈 문자열이면 'N'으로 처리
            if pick_value.strip() == "":
                return 'N'
                
            # Pick 값 검증 - P, B 외의 값이 나오면 경고 로그
            if pick_value not in ['P', 'B', 'N']:
                self.logger.warning(f"예상치 못한 PICK 값: {pick_value}, 엑셀 파일 검증 필요")
            
            return pick_value
        except Exception as e:
            self.logger.error(f"다음 열 PICK 값 확인 중 오류 발생: {e}")
            self.base_manager.reopen_excel_if_needed()
            return 'N'  # 오류 발생 시 기본값 'N' 반환
    
    def get_next_column_letter(self, column):
        """
        지정된 열의 다음 열 문자를 반환합니다.
        
        Args:
            column (str): 열 문자 (예: 'B')
            
        Returns:
            str: 다음 열 문자 (예: 'C')
        """
        try:
            col_idx = openpyxl.utils.column_index_from_string(column)
            return openpyxl.utils.get_column_letter(col_idx + 1)
        except Exception as e:
            self.logger.error(f"다음 열 문자 가져오기 실패: {e}")
            return None
    
    def get_prev_column_letter(self, column):
        """
        지정된 열의 이전 열 문자를 반환합니다.
        
        Args:
            column (str): 열 문자 (예: 'C')
            
        Returns:
            str: 이전 열 문자 (예: 'B') 또는 열이 'A'인 경우 None
        """
        try:
            col_idx = openpyxl.utils.column_index_from_string(column)
            if col_idx <= 1:  # 'A' 이하면 이전 열이 없음
                return None
            return openpyxl.utils.get_column_letter(col_idx - 1)
        except Exception as e:
            self.logger.error(f"이전 열 문자 가져오기 실패: {e}")
            return None