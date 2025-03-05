# modules/test_excel.py
"""
엑셀 파일에서 3행에 결과(P/B)를 입력하고 다음 열의 12행 PICK 값이 B나 P가 나올 때까지 테스트하는 코드
- COM 인터페이스를 사용하여 실제 Excel 프로그램으로 저장
"""
import sys
import os
import time
import random
import openpyxl
from typing import Dict, Any

# Windows에서만 사용 가능한 COM 인터페이스
try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

def find_next_column(excel_path="AUTO.xlsx"):
    """다음 빈 열 찾기 (3행 기준)"""
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    
    for col_idx in range(2, 75):  # B(2)부터 BW(75)까지
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell_value = sheet[f"{col_letter}3"].value
        
        if cell_value is None or cell_value == "":
            workbook.close()
            return col_letter
    
    workbook.close()
    return None

def get_next_column_letter(column):
    """현재 열의 다음 열 문자 반환"""
    col_idx = openpyxl.utils.column_index_from_string(column)
    return openpyxl.utils.get_column_letter(col_idx + 1)

def read_row_with_openpyxl(file_path: str, row_number: int = 12, start_col: str = 'B', end_col: str = 'R') -> Dict[str, Any]:
    """
    openpyxl을 사용하여 엑셀 파일에서 특정 행의 셀 값을 읽어옵니다.
    수식 계산 결과값을 반환합니다.
    
    Args:
        file_path (str): 엑셀 파일 경로
        row_number (int): 읽어올 행 번호 (기본값: 12)
        start_col (str): 시작 열 문자 (기본값: 'B')
        end_col (str): 끝 열 문자 (기본값: 'R')
    
    Returns:
        Dict[str, Any]: 열 이름을 키로 하고 셀 값을 값으로 하는 딕셔너리
    """
    # 엑셀 워크북 로드 (data_only=True로 설정하여 수식 대신 값을 읽어옴)
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # 활성화된 시트 (또는 첫 번째 시트) 선택
    sheet = workbook.active
    
    # 열 문자를 열 인덱스로 변환 (A=1, B=2, ...)
    start_col_idx = openpyxl.utils.column_index_from_string(start_col)
    end_col_idx = openpyxl.utils.column_index_from_string(end_col)
    
    # 결과를 저장할 딕셔너리 초기화
    result = {}
    
    # 지정된 행의 시작 열부터 끝 열까지 값을 읽어옴
    for col_idx in range(start_col_idx, end_col_idx + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell_value = sheet[f"{col_letter}{row_number}"].value
        result[col_letter] = cell_value
    
    # 워크북 닫기
    workbook.close()
    
    return result

def check_pick_value(excel_path, column):
    """12행 PICK 값 확인 (read_row_with_openpyxl 함수 사용)"""
    # 전체 12행 값 가져오기
    row_values = read_row_with_openpyxl(excel_path, 12, column, column)
    
    # 지정한 열의 값 가져오기
    pick_value = row_values.get(column)
    
    # None인 경우 "N"으로 처리
    if pick_value is None:
        return "N"
    
    # 문자열이 아닌 경우 문자열로 변환
    if not isinstance(pick_value, str):
        pick_value = str(pick_value)
    
    # 디버깅 정보 출력
    print(f"DEBUG: {column}12 셀 값 = '{pick_value}', 타입 = {type(pick_value)}")
    
    return pick_value

def write_result(excel_path, column, result):
    """3행에 결과 쓰기"""
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    sheet[f"{column}3"] = result
    workbook.save(excel_path)
    workbook.close()

def save_with_excel_app(excel_path):
    """
    실제 Excel 애플리케이션을 사용하여 파일을 열고 저장합니다.
    이 함수는 Windows에서만 작동합니다.
    """
    if not HAS_WIN32COM:
        print("win32com 라이브러리가 설치되지 않았습니다. pip install pywin32로 설치하세요.")
        return False
    
    try:
        print("Excel 애플리케이션으로 파일 열고 저장 중...")
        abs_path = os.path.abspath(excel_path)
        
        # Excel 애플리케이션 실행
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel UI를 표시하지 않음
        
        # 파일 열기
        workbook = excel.Workbooks.Open(abs_path)
        
        # 변경 없이 저장
        workbook.Save()
        
        # 파일 닫기
        workbook.Close(True)
        
        # Excel 종료
        excel.Quit()
        
        print("Excel 애플리케이션으로 파일 저장 완료")
        return True
    except Exception as e:
        print(f"Excel 애플리케이션 저장 중 오류 발생: {str(e)}")
        return False

def manual_save_instruction():
    """사용자에게 수동 저장 안내"""
    print("\n" + "=" * 50)
    print("자동 Excel 저장에 실패했습니다.")
    print("자동 테스트를 계속하려면 엑셀 파일을 수동으로 저장해주세요.")
    print("1. Excel에서 AUTO.xlsx 파일을 엽니다.")
    print("2. Ctrl+S를 눌러 저장합니다.")
    print("3. Excel을 닫습니다.")
    print("=" * 50)
    input("위 작업을 완료한 후 엔터 키를 눌러 계속하세요...")

def run_test(excel_path="AUTO.xlsx", max_attempts=50, use_excel_app=True):
    """PICK 값이 B나 P가 나올 때까지 테스트"""
    print(f"엑셀 파일: {excel_path}")
    print(f"최대 시도 횟수: {max_attempts}")
    
    # 자동 Excel 애플리케이션 사용 여부 확인
    if use_excel_app and not HAS_WIN32COM:
        print("win32com 라이브러리가 설치되지 않아 Excel 애플리케이션을 사용할 수 없습니다.")
        use_excel_app = False
    
    for attempt in range(1, max_attempts + 1):
        print(f"\n시도 #{attempt}/{max_attempts}")
        
        # 현재 빈 열 찾기
        current_column = find_next_column(excel_path)
        if not current_column:
            print("더 이상 빈 열이 없습니다.")
            return False
        
        # 다음 열 계산 (PICK 값을 확인할 열)
        next_column = get_next_column_letter(current_column)
        
        print(f"현재 입력 열: {current_column}, PICK 확인 열: {next_column}")
        
        # 랜덤 결과 생성 (P 또는 B)
        result = random.choice(['P', 'B'])
        print(f"랜덤 결과: {result}")
        
        # 현재 열에 결과 쓰기
        write_result(excel_path, current_column, result)
        print(f"{current_column}3 열에 '{result}' 입력 완료")
        
        # Excel 애플리케이션으로 저장하거나 수동 저장 안내
        if use_excel_app:
            success = save_with_excel_app(excel_path)
            if not success:
                manual_save_instruction()
        else:
            manual_save_instruction()
        
        # 다음 열의 PICK 값 확인
        pick_value = check_pick_value(excel_path, next_column)
        print(f"{next_column}12 PICK 값: {pick_value}")
        
        # B나 P이면 성공
        if pick_value in ['B', 'P']:
            print(f"\n성공! {next_column}12에서 PICK 값 '{pick_value}'를 찾았습니다!")
            
            # 전체 12행 값 확인 (결과 확인용)
            row_12_values = read_row_with_openpyxl(excel_path)
            print("\n12행 전체 값:")
            for col, val in row_12_values.items():
                if val in ['B', 'P']:
                    print(f"{col}12: {val} ← PICK 값")
                else:
                    print(f"{col}12: {val}")
            
            return True
    
    print(f"\n최대 시도 횟수({max_attempts})를 초과했습니다.")
    return False

if __name__ == "__main__":
    # 기본 파일 경로 설정
    excel_path = "AUTO.xlsx"
    
    # 명령줄에서 파일 경로를 입력받은 경우 해당 경로 사용
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    
    # 최대 시도 횟수
    max_attempts = 50
    if len(sys.argv) > 2:
        try:
            max_attempts = int(sys.argv[2])
        except ValueError:
            pass
    
    # 운영 체제 확인
    is_windows = os.name == 'nt'
    
    # Excel 애플리케이션 사용 여부 (Windows만 가능)
    use_excel_app = is_windows and HAS_WIN32COM
    
    # Excel 애플리케이션 사용 불가 시 수동 저장 안내
    if not use_excel_app:
        print("\n자동 Excel 애플리케이션 저장을 사용할 수 없습니다.")
        print("테스트 중 수동으로 엑셀 파일을 저장해야 합니다.")
        proceed = input("계속하시겠습니까? (y/n): ").strip().lower()
        if proceed != 'y':
            print("테스트를 종료합니다.")
            sys.exit(0)
    
    # 테스트 실행
    success = run_test(excel_path, max_attempts, use_excel_app)
    
    # 결과 출력
    print("\n===== 테스트 결과 =====")
    print(f"테스트 {'성공' if success else '실패'}")